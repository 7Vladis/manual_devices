from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Q, Prefetch, Count
from django.utils import timezone
from datetime import timedelta, datetime, date
from dateutil.relativedelta import relativedelta
from users.decorators import role_required
from .youtrack_services import (
    send_comment_to_youtrack,
    add_work_item_to_youtrack,
    upload_attachment_to_youtrack,
    sync_issue_from_youtrack,
    update_issue_description_in_youtrack,
    delete_comment_from_youtrack,
    delete_attachment_from_youtrack
)
from .models import DateUpdateRule, DataObject, ActionHistory, ObjectModel, ObjectType, Attachment, Comment


# --- НАСТРОЙКИ СИСТЕМЫ ---

@login_required
@role_required(['senior', 'admin', 'superuser'])
def settings_page(request):
    """Единый центр управления системой"""
    active_tab = request.GET.get('tab', 'rules')
    
    # Защита вкладок на уровне бэкенда для Старшего инженера
    if active_tab in ['notifications', 'users'] and not request.user.is_admin_or_higher:
        active_tab = 'rules'

    context = {'active_tab': active_tab}

    # 1. ВКЛАДКА: Правила планирования ТО
    if active_tab == 'rules':
        rules_query = DateUpdateRule.objects.prefetch_related('data_objects').annotate(
            objects_count=Count('data_objects')
        ).order_by('name')
        
        month_names = {
            1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля',
            5: 'Мая', 6: 'Июня', 7: 'Июля', 8: 'Августа',
            9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
        }
        
        for r in rules_query:
            rule_data = r.rule or {}
            if rule_data.get('strategy') == 'fixed':
                dates_list = rule_data.get('value', [])
                formatted_dates = []
                for d in dates_list:
                    day = d.get('day', 1)
                    month_num = d.get('month', 1)
                    month_name = month_names.get(month_num, '')
                    formatted_dates.append(f"{day} {month_name}")
                r.formatted_fixed_dates = ", ".join(formatted_dates)
                
        context['rules'] = rules_query

    # 2. ВКЛАДКА: Типы объектов
    elif active_tab == 'object_types':
        context['object_types'] = ObjectType.objects.annotate(
            models_count=Count('models')
        ).order_by('type')

    # 3. ВКЛАДКА: Уведомления Mattermost (Только для Админов и Суперпользователей)
    elif active_tab == 'notifications' and request.user.is_admin_or_higher:
        from notifications.models import MattermostSetting
        context['settings'] = MattermostSetting.objects.all().order_by('-updated_at')

    # 4. ВКЛАДКА: Пользователи (Только для Админов и Суперпользователей)
    elif active_tab == 'users' and request.user.is_admin_or_higher:
        User = get_user_model()
        context['users_list'] = User.objects.all().order_by('username')

    if request.headers.get('HX-Request'):
        return render(request, 'data/settings/settings_layout_inner.html', context)
    return render(request, 'data/settings.html', context)


@login_required
@role_required(['senior', 'admin', 'superuser'])
def create_object_type_view(request):
    """Создание нового типа оборудования из настроек"""
    if request.method == 'POST':
        name = request.POST.get('type', '').strip()
        if name:
            ObjectType.objects.get_or_create(type=name)
            
    response = HttpResponse()
    response['HX-Redirect'] = '/settings/?tab=object_types'
    return response


@login_required
@role_required(['senior', 'admin', 'superuser'])
def delete_object_type_view(request, pk):
    """Удаление типа оборудования с проверкой на использование"""
    obj_type = get_object_or_404(ObjectType, pk=pk)
    if obj_type.models.exists():
        return HttpResponse(
            '<div class="alert alert-danger py-2 px-3 m-0 rounded-3 small animate-fade">'
            '<i class="bi bi-exclamation-triangle-fill me-1"></i> '
            'Нельзя удалить тип: он привязан к существующим моделям оборудования!'
            '</div>', 
            status=400
        )
    obj_type.delete()
    
    response = HttpResponse()
    response['HX-Redirect'] = '/settings/?tab=object_types'
    return response


def parse_rule_from_request(request):
    """Разбор параметров правила планирования из POST-запроса"""
    strategy = request.POST.get('new_rule_strategy', 'relative')
    if strategy == 'relative':
        anchor = request.POST.get('new_rule_anchor', 'actual')
        try:
            years = int(request.POST.get('new_rule_years', 0))
        except (ValueError, TypeError):
            years = 0
        try:
            months = int(request.POST.get('new_rule_months', 6))
        except (ValueError, TypeError):
            months = 6
        try:
            days = int(request.POST.get('new_rule_days', 0))
        except (ValueError, TypeError):
            days = 0
            
        return {
            "strategy": "relative",
            "anchor": anchor,
            "value": {
                "years": years,
                "months": months,
                "days": days
            }
        }
    elif strategy == 'fixed':
        fixed_months = request.POST.getlist('fixed_months')
        fixed_days = request.POST.getlist('fixed_days')
        
        dates_list = []
        for m, d in zip(fixed_months, fixed_days):
            try:
                dates_list.append({"month": int(m), "day": int(d)})
            except (ValueError, TypeError):
                pass
                
        return {
            "strategy": "fixed",
            "anchor": "yearly",
            "value": dates_list
        }
    return {"strategy": "relative", "anchor": "actual", "value": {"years": 0, "months": 6, "days": 0}}


@login_required
@role_required(['senior', 'admin', 'superuser'])
def create_rule_settings_view(request):
    """Создание нового правила планирования ТО из настроек"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            rule_json = parse_rule_from_request(request)
            DateUpdateRule.objects.get_or_create(
                name=name,
                defaults={"rule": rule_json}
            )
            
        response = HttpResponse()
        response['HX-Redirect'] = '/settings/?tab=rules'
        return response

    month_names = {
        1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля',
        5: 'Мая', 6: 'Июня', 7: 'Июля', 8: 'Августа',
        9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
    }
    
    return render(request, 'data/settings/create_rule_modal.html', {
        'strategy': 'relative',
        'anchor': 'actual',
        'years': 0,
        'months': 6,
        'days': 0,
        'fixed_dates': [],
        'month_names': month_names,
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_rule_settings_view(request, pk):
    """Редактирование параметров правила планирования ТО"""
    rule_obj = get_object_or_404(DateUpdateRule, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        strategy = request.POST.get('new_rule_strategy', 'relative')
        
        if name:
            rule_obj.name = name
            
        if strategy == 'relative':
            anchor = request.POST.get('new_rule_anchor', 'actual')
            years = int(request.POST.get('new_rule_years', 0))
            months = int(request.POST.get('new_rule_months', 6))
            days = int(request.POST.get('new_rule_days', 0))
            
            rule_json = {
                "strategy": "relative",
                "anchor": anchor,
                "value": {
                    "years": years,
                    "months": months,
                    "days": days
                }
            }
        elif strategy == 'fixed':
            fixed_months = request.POST.getlist('fixed_months')
            fixed_days = request.POST.getlist('fixed_days')
            
            dates_list = []
            for m, d in zip(fixed_months, fixed_days):
                dates_list.append({"month": int(m), "day": int(d)})
                
            rule_json = {
                "strategy": "fixed",
                "anchor": "yearly",
                "value": dates_list
            }
            
        rule_obj.rule = rule_json
        rule_obj.save()
        
        # Пересчитываем даты для объектов с этим правилом
        for obj in rule_obj.data_objects.all():
            obj.next_maintenance_date = calculate_next_maintenance_date(obj, base_date=timezone.localdate())
            obj.save(update_fields=['next_maintenance_date'])
            
        response = HttpResponse()
        response['HX-Redirect'] = '/settings/?tab=rules'
        return response

    rule_data = rule_obj.rule or {}
    strategy = rule_data.get('strategy', 'relative')
    anchor = rule_data.get('anchor', 'actual')
    val = rule_data.get('value', {})
    
    years = val.get('years', 0) if strategy == 'relative' else 0
    months = val.get('months', 6) if strategy == 'relative' else 0
    days = val.get('days', 0) if strategy == 'relative' else 0
    fixed_dates = val if strategy == 'fixed' else []
    
    month_names = {
        1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля',
        5: 'Мая', 6: 'Июня', 7: 'Июля', 8: 'Августа',
        9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
    }
    
    return render(request, 'data/settings/edit_rule_modal.html', {
        'rule_obj': rule_obj,
        'strategy': strategy,
        'anchor': anchor,
        'years': years,
        'months': months,
        'days': days,
        'fixed_dates': fixed_dates,
        'month_names': month_names,
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def delete_rule_view(request, pk):
    """Удаление правила планирования"""
    rule = get_object_or_404(DateUpdateRule, pk=pk)
    if rule.data_objects.exists():
        return HttpResponse(
            '<div class="alert alert-danger py-2 px-3 m-0 rounded-3 small animate-fade">'
            '<i class="bi bi-exclamation-triangle-fill me-1"></i> '
            'Нельзя удалить: правило используется в активных объектах!'
            '</div>', 
            status=400
        )
    rule.delete()
    
    response = HttpResponse()
    response['HX-Redirect'] = '/settings/?tab=rules'
    return response


# --- ЭКСПОРТ ДАННЫХ ---

@login_required
@role_required(['admin', 'superuser'])
def export_modal_view(request):
    """Модальное окно выбора опции экспорта данных в XLSX"""
    return render(request, 'data/export_modal.html')


@login_required
@role_required(['admin', 'superuser'])
def export_xlsx_view(request):
    """Генерация XLSX файла с прямой иерархией через parent"""
    import openpyxl

    export_mode = request.GET.get('export_mode', 'all')
    
    queryset = DataObject.objects.select_related('model', 'model__object_type', 'parent', 'parent__model').all()
    if export_mode == 'with_inventory':
        queryset = queryset.filter(inventory_number__isnull=False).exclude(inventory_number='')
    
    queryset = queryset.order_by('inventory_number', 'name')

    def get_object_hierarchy_path(obj):
        """Сборка пути от корня до текущего объекта через цепочку parent"""
        path_segments = []
        current = obj
        visited = set()
        
        while current and current.uuid not in visited:
            visited.add(current.uuid)
            name = current.name or (current.model.name if current.model else "Без имени")
            path_segments.append(name)
            current = current.parent
            
        path_segments.reverse()
        return " → ".join(path_segments)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оборудование"

    headers = [
        "Инвентарный номер",
        "Название объекта",
        "Иерархический путь (от корня)",
        "Описание объекта",
        "Название модели",
        "Характеристики спецификации"
    ]
    ws.append(headers)

    for col_num in range(1, 7):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for obj in queryset:
        inv_num = obj.inventory_number or ""
        obj_name = obj.name or ""
        hierarchy_path = get_object_hierarchy_path(obj)
        description = obj.description or ""
        model_name = obj.model.name if obj.model else ""
        
        specs = obj.model.specifications if obj.model and obj.model.specifications else {}
        specs_str_list = []
        if isinstance(specs, dict):
            for k, v in specs.items():
                specs_str_list.append(f"{k}: {v}")
        specs_formatted = "; ".join(specs_str_list)

        ws.append([inv_num, obj_name, hierarchy_path, description, model_name, specs_formatted])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 15), 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"manual_devices_export_{timezone.localdate().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# --- ДАШБОРД ---

def get_period_limits(period_type):
    """Возвращает даты начала и конца периода (типа date)"""
    today = timezone.localdate()
    if period_type == 'week':
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif period_type == 'month':
        start = today.replace(day=1)
        next_month = (today.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = next_month - timedelta(days=1)
    else:
        start = end = today
    return start, end


@login_required
def dashboard(request):
    today = timezone.localdate()
    total_objects = DataObject.objects.count()
    overdue_count = DataObject.objects.filter(next_maintenance_date__lt=today).count()
    
    def get_stats(period):
        start, end = get_period_limits(period)
        completed = ActionHistory.objects.filter(
            created_at__date__range=(start, end),
            action_type='maintenance'
        ).count()
        current_planned = DataObject.objects.filter(next_maintenance_date__range=(start, end)).count()
        planned = current_planned + completed
        return completed, planned

    week_done, week_all = get_stats('week')
    month_done, month_all = get_stats('month')

    context = {
        'total_objects': total_objects,
        'overdue_count': overdue_count,
        'week_done': week_done,
        'week_all': week_all,
        'week_percent': (week_done / week_all * 100) if week_all > 0 else 0,
        'month_done': month_done,
        'month_all': month_all,
        'month_percent': (month_done / month_all * 100) if month_all > 0 else 0,
    }
    return render(request, 'data/dashboard.html', context)


@login_required
def maintenance_list(request):
    period = request.GET.get('period', 'week')
    today = timezone.localdate()
    
    if period == 'overdue':
        objects = DataObject.objects.filter(next_maintenance_date__lt=today)
        label = "Просроченные ТО"
    elif period == 'all':
        objects = DataObject.objects.all()
        label = "Все объекты системы"
    elif period == 'month':
        _, end = get_period_limits('month')
        objects = DataObject.objects.filter(next_maintenance_date__lte=end)
        label = "План на месяц"
    else:
        _, end = get_period_limits('week')
        objects = DataObject.objects.filter(next_maintenance_date__lte=end)
        label = "План на неделю"
        
    objects = objects.select_related('model', 'model__object_type').order_by('next_maintenance_date')

    return render(request, 'data/includes/maintenance_table.html', {
        'objects': objects,
        'today': today,
        'period_label': label
    })


@login_required
def search_view(request):
    query = request.GET.get('q', '').strip()
    if not query or len(query) < 2:
        return HttpResponse('')

    base_filters = (
        Q(name__icontains=query) |
        Q(inventory_number__icontains=query) |
        Q(description__icontains=query) |
        Q(model__name__icontains=query) |
        Q(model__specifications__icontains=query) |
        Q(comments__text__icontains=query) |
        Q(actions__action__icontains=query)
    )
    
    results = DataObject.objects.filter(base_filters).distinct()
    
    if results.count() < 3:
        words = query.split()
        if len(words) > 1:
            word_filters = Q()
            for word in words:
                word_filters |= (
                    Q(name__icontains=word) | 
                    Q(description__icontains=word) | 
                    Q(model__name__icontains=word) |
                    Q(model__specifications__icontains=word)
                )
            results = (results | DataObject.objects.filter(word_filters)).distinct()

    return render(request, 'data/includes/search_results_list.html', {'results': results[:10]})


# --- СПРАВОЧНИК И ПРОВОДНИК ---

@login_required
def dict_view(request):
    active_tab = request.GET.get('tab', 'objects')
    
    selected_object_id = request.GET.get('object')
    selected_model_id = request.GET.get('model')
    
    active_object = None
    active_model = None
    parent_uuids = []
    
    explorer_mode = request.session.get('explorer_mode', 'tree')
    explorer_parent_uuid = request.session.get('explorer_parent_uuid')
    explorer_parent = None
    
    if selected_object_id:
        active_tab = 'objects'
        active_object = get_object_or_404(DataObject.objects.select_related('parent'), pk=selected_object_id)
        
        current = active_object.parent
        while current:
            parent_uuids.append(str(current.pk))
            current = current.parent
                
        if explorer_mode == 'flat':
            if active_object.parent:
                explorer_parent_uuid = str(active_object.parent.pk)
                request.session['explorer_parent_uuid'] = explorer_parent_uuid
            else:
                explorer_parent_uuid = None
                request.session['explorer_parent_uuid'] = None
                
    elif selected_model_id:
        active_tab = 'models'
        active_model = get_object_or_404(ObjectModel, pk=selected_model_id)
        
    if explorer_mode == 'flat' and explorer_parent_uuid:
        try:
            explorer_parent = DataObject.objects.get(pk=explorer_parent_uuid)
        except DataObject.DoesNotExist:
            explorer_parent_uuid = None
            request.session['explorer_parent_uuid'] = None
    
    models = ObjectModel.objects.all().order_by('name')
    object_types = ObjectType.objects.all().order_by('type')
    
    context = {
        'models': models,
        'object_types': object_types,
        'active_tab': active_tab,
        'active_object': active_object,
        'active_model': active_model,
        'parent_uuids': parent_uuids,
        'explorer_mode': explorer_mode,
        'explorer_parent': explorer_parent,
    }
    
    if active_tab == 'models':
        context['object_types_list'] = ObjectType.objects.prefetch_related(
            Prefetch('models', queryset=ObjectModel.objects.all().order_by('name'))
        ).order_by('type')
    else:
        if explorer_mode == 'flat' and explorer_parent:
            context['initial_objects'] = explorer_parent.children.all().prefetch_related('children').order_by('name')
        else:
            context['initial_objects'] = DataObject.objects.filter(
                parent__isnull=True
            ).prefetch_related('children').order_by('name')
        
    if request.headers.get('HX-Request') and request.GET.get('sidebar'):
        return render(request, 'data/tree/dict_sidebar.html', context)
        
    return render(request, 'data/dict.html', context)


@login_required
def toggle_explorer_mode_view(request):
    """Переключает режим отображения проводника"""
    if request.method == 'POST':
        current_mode = request.session.get('explorer_mode', 'tree')
        new_mode = 'flat' if current_mode == 'tree' else 'tree'
        request.session['explorer_mode'] = new_mode
        request.session['explorer_parent_uuid'] = None
        
        mode_display = "Проводник" if new_mode == 'flat' else "Дерево"
        checked_attr = "checked" if new_mode == 'flat' else ""
        
        html = f"""
        <div id="explorer-toggle-wrapper" class="d-flex align-items-center justify-content-between w-100" style="min-height: 40px;">
            <div class="d-flex flex-column text-start" style="user-select: none;">
                <small class="text-muted fw-bold text-uppercase" style="font-size: 0.65rem; letter-spacing: 0.5px; line-height: 1.2;">
                    Режим справочника
                </small>
                <span class="text-primary fw-bold" id="explorer-mode-text" style="font-size: 0.85rem; line-height: 1.2; margin-top: 2px;">
                    {mode_display}
                </span>
            </div>
            <div class="form-check form-switch m-0 d-flex align-items-center">
                <input class="form-check-input" 
                       type="checkbox" 
                       id="explorerModeToggle"
                       hx-post="/dict/toggle-explorer-mode/"
                       hx-target="#explorer-toggle-wrapper"
                       hx-swap="outerHTML"
                       style="cursor: pointer; transform: scale(1.1); margin: 0;"
                       {checked_attr}>
            </div>
        </div>
        """
        
        response = HttpResponse(html)
        response['HX-Trigger'] = 'explorerModeChanged'
        return response
        
    return HttpResponse("Метод не разрешен", status=405)


@login_required
def explorer_navigate_view(request, pk):
    """Переход внутрь объекта в плоском режиме"""
    request.session['explorer_parent_uuid'] = str(pk)
    response = HttpResponse()
    response['HX-Trigger'] = 'explorerModeChanged'
    return response


@login_required
def explorer_up_view(request):
    """Переход на один уровень вверх в плоском режиме"""
    parent_uuid = request.session.get('explorer_parent_uuid')
    if parent_uuid:
        try:
            current_parent = DataObject.objects.select_related('parent').get(pk=parent_uuid)
            request.session['explorer_parent_uuid'] = str(current_parent.parent.pk) if current_parent.parent else None
        except DataObject.DoesNotExist:
            request.session['explorer_parent_uuid'] = None
            
    response = HttpResponse()
    response['HX-Trigger'] = 'explorerModeChanged'
    return response


@login_required
def object_tree_view(request):
    roots = DataObject.objects.filter(parent__isnull=True).prefetch_related('children').order_by('name')
    return render(request, 'data/tree/object_tree_list.html', {'objects': roots})


@login_required
def object_children_view(request, parent_uuid):
    parent = get_object_or_404(DataObject, pk=parent_uuid)
    children = parent.children.all().prefetch_related('children').order_by('name')
    
    active_object_id = request.GET.get('active_object')
    active_object = None
    parent_uuids = []
    
    if active_object_id:
        try:
            active_object = DataObject.objects.select_related('parent').get(pk=active_object_id)
            current = active_object.parent
            while current:
                parent_uuids.append(str(current.pk))
                current = current.parent
        except DataObject.DoesNotExist:
            pass

    return render(request, 'data/tree/object_tree_list_nodes.html', {
        'objects': children,
        'parent': parent,
        'active_object': active_object,
        'parent_uuids': parent_uuids,
    })


@login_required
def model_tree_view(request):
    object_types = ObjectType.objects.prefetch_related(
        Prefetch('models', queryset=ObjectModel.objects.all().order_by('name'))
    ).order_by('type')
    return render(request, 'data/tree/model_tree_list.html', {'object_types': object_types})


# --- ОБСЛУЖИВАНИЕ ОБЪЕКТА ---

@login_required
def service_object_view(request, pk):
    """Выполнение ТО объекта"""
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'POST':
        date_str = request.POST.get('maintenance_date')
        spent_time = request.POST.get('spent_time', '').strip()
        custom_comment = request.POST.get('comment', '').strip()
        
        if date_str:
            maintenance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            obj.next_maintenance_date = maintenance_date
            obj.save(update_fields=['next_maintenance_date'])
            
            action_text = custom_comment if custom_comment else "Плановое техническое обслуживание выполнено"
            
            history_entry = ActionHistory.objects.create(
                user=request.user,
                data_object=obj,
                action_type='maintenance',
                action=action_text
            )

            if obj.youtrack_issue_id and spent_time:
                work_item_desc = custom_comment if custom_comment else f"Техническое обслуживание (след. ТО: {maintenance_date.strftime('%d.%m.%Y')})"
                
                success_work, work_id_or_err = add_work_item_to_youtrack(
                    issue_id=obj.youtrack_issue_id,
                    duration_str=spent_time,
                    text=work_item_desc,
                    user=request.user
                )
                
                if success_work and work_id_or_err:
                    history_entry.youtrack_id = work_id_or_err
                    history_entry.save(update_fields=['youtrack_id'])
        
        node_html = render_to_string('data/tree/object_tree_node.html', {'node': obj}, request=request)
        date_display_str = obj.next_maintenance_date.strftime('%d.%m.%Y') if obj.next_maintenance_date else "Не запланировано"
        oob_date_html = f'<strong id="maintenance-date-display" class="text-danger" hx-swap-oob="innerHTML">{date_display_str}</strong>'
        
        response = HttpResponse(node_html + "\n" + oob_date_html)
        response['HX-Trigger'] = 'objectServiced'
        return response

    proposed_date = calculate_next_maintenance_date(obj)
    
    return render(request, 'data/tree/service_modal_body.html', {
        'obj': obj,
        'proposed_date': proposed_date
    })


# --- УДАЛЕНИЕ ---

@login_required
@role_required(['senior', 'admin', 'superuser'])
def delete_object_view(request, pk):
    if request.method in ['POST', 'DELETE']:
        obj = get_object_or_404(DataObject, pk=pk)
        obj.delete()
        return HttpResponse("", status=200)
    return HttpResponse("Метод не разрешен", status=405)


@login_required
@role_required(['senior', 'admin', 'superuser'])
def delete_model_view(request, pk):
    if request.method in ['POST', 'DELETE']:
        model_obj = get_object_or_404(ObjectModel, pk=pk)
        model_obj.delete()
        return HttpResponse("", status=200)
    return HttpResponse("Метод не разрешен", status=405)


# --- СОЗДАНИЕ ОБЪЕКТОВ И МОДЕЛЕЙ ---

@login_required
@role_required(['senior', 'admin', 'superuser'])
def create_object_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        model_uuid = request.POST.get('model')
        inventory_number = request.POST.get('inventory_number')
        youtrack_issue_id = request.POST.get('youtrack_issue_id', '').strip()
        parent_uuid = request.POST.get('parent')
        maintenance_str = request.POST.get('next_maintenance_date')
        rule_uuid = request.POST.get('date_update_rule')
        
        model_obj = get_object_or_404(ObjectModel, pk=model_uuid)
        
        next_maintenance_date = None
        if maintenance_str:
            try:
                next_maintenance_date = datetime.strptime(maintenance_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        selected_rule = get_object_or_404(DateUpdateRule, pk=rule_uuid) if rule_uuid else None
        parent_obj = get_object_or_404(DataObject, pk=parent_uuid) if parent_uuid else None
                
        new_obj = DataObject.objects.create(
            name=name,
            model=model_obj,
            parent=parent_obj,
            inventory_number=inventory_number if inventory_number else None,
            youtrack_issue_id=youtrack_issue_id if youtrack_issue_id else None,
            next_maintenance_date=next_maintenance_date,
            date_update_rule=selected_rule
        )

        scheduling_mode = request.POST.get('maintenance_scheduling_mode', 'manual')
        if scheduling_mode == 'auto' and selected_rule:
            first_date = calculate_next_maintenance_date(new_obj, base_date=timezone.localdate())
            new_obj.next_maintenance_date = first_date
            new_obj.save(update_fields=['next_maintenance_date'])
            
        if parent_obj:
            parent_name = parent_obj.name or parent_obj.model.name
            action_desc = f"Объект зарегистрирован в системе в составе родительского объекта '{parent_name}'."
        else:
            action_desc = "Объект зарегистрирован в системе как корневой объект."

        ActionHistory.objects.create(
            user=request.user,
            data_object=new_obj,
            action_type='create',
            action=action_desc
        )
        
        roots = DataObject.objects.filter(parent__isnull=True).order_by('name')
        context = {
            'initial_objects': roots,
            'active_tab': 'objects',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        return render(request, 'data/tree/dict_sidebar.html', context)

    return render(request, 'data/includes/create_object_modal_body.html')


@login_required
@role_required(['senior', 'admin', 'superuser'])
def create_model_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        type_uuid = request.POST.get('object_type')
        new_type_name = request.POST.get('new_object_type')
        
        spec_keys = request.POST.getlist('spec_keys')
        spec_values = request.POST.getlist('spec_values')
        specifications = dict(zip(spec_keys, spec_values))
        
        if new_type_name:
            object_type, _ = ObjectType.objects.get_or_create(type=new_type_name)
        else:
            object_type = get_object_or_404(ObjectType, pk=type_uuid)
            
        ObjectModel.objects.create(
            name=name,
            object_type=object_type,
            specifications=specifications
        )
        
        object_types = ObjectType.objects.prefetch_related(
            Prefetch('models', queryset=ObjectModel.objects.all().order_by('name'))
        ).order_by('type')
        context = {
            'object_types_list': object_types,
            'active_tab': 'models',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        return render(request, 'data/tree/dict_sidebar.html', context)

    return render(request, 'data/includes/create_model_modal_body.html')


# --- ДЕТАЛИ ОБЪЕКТА ---

@login_required
def object_detail_view(request, pk):
    obj = get_object_or_404(
        DataObject.objects.select_related('model', 'model__object_type', 'parent', 'parent__model'), 
        pk=pk
    )
    prev_active_id = request.session.get('active_object_id')
    request.session['active_object_id'] = str(pk)
    
    context = {
        'obj': obj,
        'parent': obj.parent,
        'active_tab': 'short_info',
    }
    
    response_content = render_to_string('data/object/object_details.html', context, request=request)
    oob_elements = []
    
    new_node_html = render_to_string('data/tree/object_tree_node_label.html', {
        'node': obj,
        'is_active': True,
        'oob': True
    }, request=request)
    oob_elements.append(new_node_html)
    
    if prev_active_id and prev_active_id != str(pk):
        try:
            prev_obj = DataObject.objects.get(pk=prev_active_id)
            old_node_html = render_to_string('data/tree/object_tree_node_label.html', {
                'node': prev_obj,
                'is_active': False,
                'oob': True
            }, request=request)
            oob_elements.append(old_node_html)
        except DataObject.DoesNotExist:
            pass
            
    combined_content = response_content + "\n" + "\n".join(oob_elements)
    
    if request.GET.get('sidebar'):
        roots = DataObject.objects.filter(parent__isnull=True).order_by('name')
        sidebar_context = {
            'initial_objects': roots,
            'active_tab': 'objects',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        sidebar_html = render_to_string('data/tree/dict_sidebar.html', sidebar_context, request=request)
        combined_content = combined_content + f'\n<div id="sidebar-container" hx-swap-oob="innerHTML">{sidebar_html}</div>'
        
    return HttpResponse(combined_content)


@login_required
@role_required(['senior', 'admin', 'superuser'])
def unlink_rule_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    obj.date_update_rule = None
    obj.save(update_fields=['date_update_rule'])
    
    ActionHistory.objects.create(
        user=request.user,
        data_object=obj,
        action_type='rule_change',
        action="Правило автоматического расчета ТО отвязано от объекта."
    )
    
    return render(request, 'data/object/edit_rule_modal_body.html', {
        'obj': obj,
        'current_mode': 'auto',
        'rule_details': None
    })


@login_required
def object_tab_view(request, pk, tab_name):
    obj = get_object_or_404(DataObject, pk=pk)
    context = {'obj': obj}
    
    if tab_name == 'short_info':
        preview = Attachment.objects.filter(data_object=obj, is_preview=True).first()
        context['preview'] = preview
        template = 'data/object/object_tab_short_info.html'
        
    elif tab_name == 'specs':
        context['specifications'] = obj.model.specifications or {}
        template = 'data/object/object_tab_specs.html'
        
    elif tab_name == 'comments':
        comments = obj.comments.select_related('user').order_by('-created_at')
        context['comments'] = comments
        template = 'data/object/object_tab_comments.html'
        
    elif tab_name == 'files':
        files = obj.attachments.select_related('user').order_by('-created_at')
        context['files'] = files
        template = 'data/object/object_tab_files.html'
        
    elif tab_name == 'history':
        history = obj.actions.select_related('user').order_by('-created_at')
        context['history'] = history
        template = 'data/object/object_tab_history.html'
        
    else:
        return HttpResponse("Вкладка не найдена", status=404)
        
    return render(request, template, context)


# --- INLINE-РЕДАКТИРОВАНИЕ ---

@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_inventory_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    if request.method == 'POST':
        new_inv = request.POST.get('inventory_number', '').strip()
        obj.inventory_number = new_inv if new_inv else None
        obj.save(update_fields=['inventory_number'])
        
        ActionHistory.objects.create(
            user=request.user,
            data_object=obj,
            action_type='update',
            action=f"Изменен инвентарный номер объекта на: {new_inv or 'отсутствует'}."
        )
        return render(request, 'data/object/inline_inventory.html', {'obj': obj, 'editing': False})
        
    if request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_inventory.html', {'obj': obj, 'editing': False})
        
    return render(request, 'data/object/inline_inventory.html', {'obj': obj, 'editing': True})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_youtrack_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'POST':
        new_yt = request.POST.get('youtrack_issue_id', '').strip()
        obj.youtrack_issue_id = new_yt if new_yt else None
        obj.save(update_fields=['youtrack_issue_id'])
        
        ActionHistory.objects.create(
            user=request.user,
            data_object=obj,
            action_type='update',
            action=f"Изменен ID задачи Youtrack на: {new_yt or 'отсутствует'}."
        )
        return render(request, 'data/object/inline_youtrack.html', {'obj': obj, 'editing': False})
        
    if request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_youtrack.html', {'obj': obj, 'editing': False})
        
    return render(request, 'data/object/inline_youtrack.html', {'obj': obj, 'editing': True})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_parent_view(request, pk):
    """Смена родительского объекта напрямую через ForeignKey parent"""
    obj = get_object_or_404(DataObject.objects.select_related('parent', 'parent__model'), pk=pk)
    
    if request.method == 'GET' and request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_parent.html', {
            'obj': obj, 
            'current_parent': obj.parent,
            'editing': False
        })
        
    if request.method == 'POST':
        parent_uuid = request.POST.get('parent') or request.POST.get('parent_uuid')
        
        if parent_uuid:
            # Защита от установки самого себя родителем
            if str(obj.pk) == str(parent_uuid):
                return HttpResponse("Объект не может быть родителем самого себя", status=400)
            new_parent = get_object_or_404(DataObject, pk=parent_uuid)
            obj.parent = new_parent
            parent_name = new_parent.name or new_parent.model.name
        else:
            obj.parent = None
            parent_name = "отсутствует"
            
        obj.save(update_fields=['parent'])
            
        ActionHistory.objects.create(
            user=request.user,
            data_object=obj,
            action_type='link_change',
            action=f"Назначен новый родительский объект: '{parent_name}'."
        )
        
        roots = DataObject.objects.filter(parent__isnull=True).order_by('name')
        sidebar_context = {
            'initial_objects': roots,
            'active_tab': 'objects',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        sidebar_html = render(request, 'data/tree/dict_sidebar.html', sidebar_context).content.decode('utf-8')
        parent_html = render_to_string('data/object/inline_parent.html', {
            'obj': obj, 
            'current_parent': obj.parent, 
            'editing': False
        }, request=request)
            
        response_html = f"""
            {parent_html}
            <div id="sidebar-container" hx-swap-oob="innerHTML">
                {sidebar_html}
            </div>
        """
        return HttpResponse(response_html)

    return render(request, 'data/object/inline_parent.html', {
        'obj': obj,
        'current_parent': obj.parent,
        'editing': True
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_name_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'GET' and request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_name.html', {'obj': obj, 'editing': False})
        
    if request.method == 'POST':
        old_name = obj.name or obj.model.name
        new_name = request.POST.get('name', '').strip()
        
        if new_name and old_name != new_name:
            obj.name = new_name
            obj.save(update_fields=['name'])
        
            ActionHistory.objects.create(
                user=request.user,
                data_object=obj,
                action_type='update',
                action=f"Имя объекта изменено с '{old_name}' на '{new_name}'."
            )
            
        name_html = render_to_string('data/object/inline_name.html', {'obj': obj, 'editing': False}, request=request)
        sidebar_node_html = render_to_string('data/tree/object_tree_node_label.html', {
            'node': obj,
            'is_active': True,
            'oob': True
        }, request=request)
        
        return HttpResponse(name_html + "\n" + sidebar_node_html)

    return render(request, 'data/object/inline_name.html', {'obj': obj, 'editing': True})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_description_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)

    if request.method == 'GET' and request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_description.html', {'obj': obj, 'editing': False})

    if request.method == 'POST':
        old_desc = (obj.description or '').strip()
        new_desc = request.POST.get('description', '').strip()
        
        if old_desc != new_desc:
            obj.description = new_desc if new_desc else None
            obj.save(update_fields=['description'])
            
            ActionHistory.objects.create(
                user=request.user,
                data_object=obj,
                action_type='update',
                action="Обновлено описание объекта."
            )

            if obj.youtrack_issue_id:
                update_issue_description_in_youtrack(
                    issue_id=obj.youtrack_issue_id,
                    description=new_desc,
                    user=request.user
                )

        return render(request, 'data/object/inline_description.html', {'obj': obj, 'editing': False})
        
    return render(request, 'data/object/inline_description.html', {'obj': obj, 'editing': True})


# --- КОММЕНТАРИИ И ВЛОЖЕНИЯ ---

@login_required
def add_comment_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    text = request.POST.get('text', '').strip()
    file = request.FILES.get('file')
    yt_errors = []
    
    if text or file:
        comment = Comment.objects.create(
            user=request.user,
            data_object=obj,
            text=text or "Вложение к объекту"
        )
        
        attachment_obj = None
        if file:
            attachment_obj = Attachment.objects.create(
                user=request.user,
                data_object=obj,
                comment=comment,
                path=file,
                is_preview=False
            )
            
        if obj.youtrack_issue_id:
            if attachment_obj and attachment_obj.path:
                try:
                    with open(attachment_obj.path.path, 'rb') as f:
                        success_file, result_file = upload_attachment_to_youtrack(
                            issue_id=obj.youtrack_issue_id,
                            file_obj=f,
                            user=request.user
                        )
                        if success_file and result_file:
                            attachment_obj.youtrack_id = result_file
                            attachment_obj.save(update_fields=['youtrack_id'])
                        elif not success_file:
                            yt_errors.append(result_file)
                except Exception as e:
                    yt_errors.append(f"Ошибка чтения файла для YouTrack: {e}")

            yt_text = text
            if attachment_obj and attachment_obj.is_image:
                image_md = f"\n\n![]({attachment_obj.filename})"
                yt_text = (yt_text + image_md) if yt_text else f"![]({attachment_obj.filename})"
            elif not yt_text and attachment_obj:
                yt_text = f"Прикреплен файл: {attachment_obj.filename}"

            if yt_text:
                success_txt, result_txt = send_comment_to_youtrack(
                    issue_id=obj.youtrack_issue_id,
                    text=yt_text,
                    user=request.user
                )
                if success_txt and result_txt:
                    comment.youtrack_id = result_txt
                    comment.save(update_fields=['youtrack_id'])
                elif not success_txt:
                    yt_errors.append(result_txt)

    comments = obj.comments.select_related('user').prefetch_related('attachments').order_by('-created_at')
    return render(request, 'data/object/object_tab_comments.html', {
        'obj': obj, 
        'comments': comments,
        'yt_error': " | ".join(yt_errors) if yt_errors else None
    })


@login_required
def edit_comment_view(request, pk):
    comment = get_object_or_404(Comment, pk=pk)
    
    if comment.user != request.user and not request.user.is_admin_or_higher:
        return HttpResponseForbidden("Вы не можете редактировать чужие комментарии.")
        
    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        if text:
            comment.text = text
            comment.save(update_fields=['text'])
        return render(request, 'data/object/comment_item.html', {'comment': comment})
        
    return render(request, 'data/object/comment_item_edit.html', {'comment': comment})


@login_required
def delete_comments_bulk(request):
    comment_ids = request.POST.getlist('comment_ids')
    obj_pk = request.POST.get('object_uuid')
    obj = get_object_or_404(DataObject, pk=obj_pk)
    
    if comment_ids:
        queryset = Comment.objects.filter(uuid__in=comment_ids, data_object=obj).prefetch_related('attachments')
        
        if not request.user.can_manage_content:
            queryset = queryset.filter(user=request.user)
            
        if obj.youtrack_issue_id:
            for comment in queryset:
                for att in comment.attachments.all():
                    if att.youtrack_id:
                        delete_attachment_from_youtrack(
                            issue_id=obj.youtrack_issue_id,
                            attachment_yt_id=att.youtrack_id,
                            user=request.user
                        )
                if comment.youtrack_id:
                    delete_comment_from_youtrack(
                        issue_id=obj.youtrack_issue_id,
                        comment_yt_id=comment.youtrack_id,
                        user=request.user
                    )
        
        queryset.delete()
        
    comments = obj.comments.select_related('user').prefetch_related('attachments').order_by('-created_at')
    return render(request, 'data/object/object_tab_comments.html', {'obj': obj, 'comments': comments})


@login_required
def add_attachment_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    file = request.FILES.get('file')
    is_preview_upload = request.POST.get('is_preview') == 'true'
    
    if file:
        if is_preview_upload:
            content_type = getattr(file, 'content_type', '')
            filename_lower = file.name.lower()
            image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg')
            is_image_mime = content_type.startswith('image/')
            is_image_ext = filename_lower.endswith(image_extensions)
            
            if not (is_image_mime or is_image_ext):
                return HttpResponse(
                    '<div class="alert alert-danger py-2 px-3 mb-3 rounded-3 small animate-fade">'
                    '<i class="bi bi-exclamation-triangle-fill me-1"></i> '
                    'Ошибка: файл превью должен быть изображением!'
                    '</div>',
                    status=400
                )
            
            Attachment.objects.filter(data_object=obj, is_preview=True).update(is_preview=False)
            
        Attachment.objects.create(
            user=request.user,
            data_object=obj,
            path=file,
            is_preview=is_preview_upload
        )
        
    if is_preview_upload:
        preview = Attachment.objects.filter(data_object=obj, is_preview=True).first()
        return render(request, 'data/object/object_tab_short_info.html', {'obj': obj, 'preview': preview})
        
    files = obj.attachments.select_related('user').order_by('-created_at')
    return render(request, 'data/object/object_tab_files.html', {'obj': obj, 'files': files})


@login_required
def delete_attachments_bulk(request):
    file_ids = request.POST.getlist('file_ids')
    obj_pk = request.POST.get('object_uuid')
    obj = get_object_or_404(DataObject, pk=obj_pk)
    
    if file_ids:
        queryset = Attachment.objects.filter(uuid__in=file_ids, data_object=obj)
        
        if not request.user.can_manage_content:
            queryset = queryset.filter(user=request.user)
            
        if obj.youtrack_issue_id:
            for att in queryset:
                if att.youtrack_id:
                    delete_attachment_from_youtrack(
                        issue_id=obj.youtrack_issue_id,
                        attachment_yt_id=att.youtrack_id,
                        user=request.user
                    )
                    
        queryset.delete()
        
    files = obj.attachments.select_related('user').order_by('-created_at')
    return render(request, 'data/object/object_tab_files.html', {'obj': obj, 'files': files})


# --- ДЕТАЛИ МОДЕЛИ ---

@login_required
def model_detail_view(request, pk):
    model_obj = get_object_or_404(ObjectModel.objects.select_related('object_type'), pk=pk)
    prev_active_id = request.session.get('active_model_id')
    request.session['active_model_id'] = str(pk)
    
    context = {
        'model_obj': model_obj,
        'active_tab': 'specs',
    }
    
    response_content = render_to_string('data/model/model_details.html', context, request=request)
    oob_elements = []
    
    new_node_html = render_to_string('data/tree/model_tree_node_label.html', {
        'model': model_obj,
        'is_active': True,
        'oob': True
    }, request=request)
    oob_elements.append(new_node_html)
    
    if prev_active_id and prev_active_id != str(pk):
        try:
            prev_model = ObjectModel.objects.get(pk=prev_active_id)
            old_node_html = render_to_string('data/tree/model_tree_node_label.html', {
                'model': prev_model,
                'is_active': False,
                'oob': True
            }, request=request)
            oob_elements.append(old_node_html)
        except ObjectModel.DoesNotExist:
            pass
            
    combined_content = response_content + "\n" + "\n".join(oob_elements)
    
    if request.GET.get('sidebar'):
        object_types = ObjectType.objects.prefetch_related(
            Prefetch('models', queryset=ObjectModel.objects.all().order_by('name'))
        ).order_by('type')
        sidebar_context = {
            'object_types_list': object_types,
            'active_tab': 'models',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        sidebar_html = render_to_string('data/tree/dict_sidebar.html', sidebar_context, request=request)
        combined_content = combined_content + f'\n<div id="sidebar-container" hx-swap-oob="innerHTML">{sidebar_html}</div>'
        
    return HttpResponse(combined_content)


@login_required
def model_tab_view(request, pk, tab_name):
    model_obj = get_object_or_404(ObjectModel, pk=pk)
    context = {'model_obj': model_obj}
    
    if tab_name == 'specs':
        context['specifications'] = model_obj.specifications or {}
        template = 'data/model/model_tab_specs.html'
    elif tab_name == 'objects':
        objects = model_obj.data_objects.all().select_related('model__object_type').order_by('name')
        context['objects'] = objects
        template = 'data/model/model_tab_objects.html'
    else:
        return HttpResponse("Вкладка не найдена", status=404)
        
    return render(request, template, context)


# --- СПЕЦИФИКАЦИИ МОДЕЛИ ---

@login_required
@role_required(['senior', 'admin', 'superuser'])
def model_spec_add_view(request, pk):
    model_obj = get_object_or_404(ObjectModel, pk=pk)
    if request.method == 'POST':
        key = request.POST.get('key', '').strip()
        value = request.POST.get('value', '').strip()
        if key and value:
            specs = model_obj.specifications or {}
            specs[key] = value
            model_obj.specifications = specs
            model_obj.save(update_fields=['specifications'])
            
        context = {
            'model_obj': model_obj,
            'specifications': model_obj.specifications
        }
        return render(request, 'data/model/model_tab_specs.html', context)


@login_required
@role_required(['senior', 'admin', 'superuser'])
def model_spec_edit_view(request, pk):
    model_obj = get_object_or_404(ObjectModel, pk=pk)
    key = request.GET.get('key') or request.POST.get('key')
    
    if request.method == 'POST':
        old_key = request.POST.get('old_key')
        new_key = request.POST.get('key', '').strip()
        value = request.POST.get('value', '').strip()
        specs = model_obj.specifications or {}
        
        if old_key and new_key and value:
            if old_key != new_key:
                specs.pop(old_key, None)
            specs[new_key] = value
            model_obj.specifications = specs
            model_obj.save(update_fields=['specifications'])
            
        context = {
            'model_obj': model_obj,
            'specifications': model_obj.specifications
        }
        return render(request, 'data/model/model_tab_specs.html', context)
        
    value = model_obj.specifications.get(key, '')
    return render(request, 'data/model/inline_spec_row.html', {
        'model_obj': model_obj,
        'key': key,
        'value': value,
        'editing': True
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def model_spec_delete_view(request, pk):
    model_obj = get_object_or_404(ObjectModel, pk=pk)
    key = request.POST.get('key')
    specs = model_obj.specifications or {}
    
    if key in specs:
        specs.pop(key)
        model_obj.specifications = specs
        model_obj.save(update_fields=['specifications'])
        
    context = {
        'model_obj': model_obj,
        'specifications': model_obj.specifications
    }
    return render(request, 'data/model/model_tab_specs.html', context)


# --- ПОДСКАЗКИ И ПРОВЕРКИ ИМЕН ---

@login_required
def check_model_name_view(request):
    name = request.GET.get('name', '').strip()
    if not name or len(name) < 2:
        return HttpResponse('')

    exact_match = ObjectModel.objects.filter(name__iexact=name).first()
    if exact_match:
        return HttpResponse(
            f'<div class="alert alert-danger py-2 px-3 mt-2 mb-0 rounded-3 small animate-fade">'
            f'<div class="fw-bold mb-1"><i class="bi bi-x-circle-fill me-1"></i> Модель с таким названием уже существует!</div>'
            f'<a href="#" class="text-danger fw-bold" '
            f'hx-get="/dict/models/{exact_match.uuid}/?sidebar=1" '
            f'hx-target="#detail-container" '
            f'hx-on:click="bootstrap.Modal.getInstance(document.getElementById(\'createModelModal\')).hide();">'
            f'Перейти к существующей модели: {exact_match.name} ({exact_match.object_type.type})'
            f'</a>'
            f'</div>'
        )

    stop_words = {'в', 'на', 'под', 'над', 'для', 'из', 'со', 'и', 'или', 'а', 'но', 'с', 'по', 'of', 'and', 'the'}
    words = [w.lower() for w in name.split() if len(w) >= 2 and w.lower() not in stop_words]

    similar_models = []
    if words:
        query = Q()
        for word in words:
            query |= Q(name__icontains=word) | Q(object_type__type__icontains=word)
        similar_models = ObjectModel.objects.filter(query).select_related('object_type').distinct()[:5]

    if similar_models:
        links = [
            f'<li class="mb-1">'
            f'<a href="#" class="alert-link text-primary fw-semibold" '
            f'hx-get="/dict/models/{model.uuid}/?sidebar=1" '
            f'hx-target="#detail-container" '
            f'hx-on:click="bootstrap.Modal.getInstance(document.getElementById(\'createModelModal\')).hide();">'
            f'{model.name} <span class="text-muted fw-normal">({model.object_type.type})</span>'
            f'</a></li>'
            for model in similar_models
        ]
        return HttpResponse(
            f'<div class="alert alert-warning py-2 px-3 mt-2 mb-0 rounded-3 small animate-fade">'
            f'<div class="fw-bold text-dark mb-1"><i class="bi bi-exclamation-triangle-fill text-warning me-1"></i> Похожие модели ({len(similar_models)} шт.):</div>'
            f'<ul class="ps-3 mb-1" style="max-height: 80px; overflow-y: auto;">{"".join(links)}</ul>'
            f'</div>'
        )

    return HttpResponse('<div class="text-success small mt-1"><i class="bi bi-check-circle-fill me-1"></i> Название свободно</div>')


@login_required
def check_object_name_view(request):
    name = request.GET.get('name', '').strip()
    if not name or len(name) < 2:
        return HttpResponse('')

    exact_match = DataObject.objects.filter(name__iexact=name).first()
    if exact_match:
        exact_name = exact_match.name or exact_match.model.name
        return HttpResponse(
            f'<div class="alert alert-danger py-2 px-3 mt-2 mb-0 rounded-3 small">'
            f'<div class="fw-bold mb-1"><i class="bi bi-x-circle-fill me-1"></i> Объект с таким именем уже существует!</div>'
            f'<a href="#" class="text-danger fw-bold" '
            f'hx-get="/dict/objects/{exact_match.uuid}/?sidebar=1" '
            f'hx-target="#detail-container" '
            f'hx-on:click="bootstrap.Modal.getInstance(document.getElementById(\'createObjectModal\')).hide();">'
            f'Перейти к объекту: {exact_name} ({exact_match.model.name})'
            f'</a>'
            f'</div>'
        )

    stop_words = {'в', 'на', 'под', 'над', 'для', 'из', 'со', 'и', 'или', 'а', 'но', 'с', 'по'}
    words = [w.lower() for w in name.split() if len(w) >= 3 and w.lower() not in stop_words]

    similar_objects = []
    if words:
        query = Q()
        for word in words:
            query |= Q(name__icontains=word) | Q(model__name__icontains=word)
        similar_objects = DataObject.objects.filter(query).select_related('model').distinct()[:5]

    if similar_objects:
        links = [
            f'<li class="mb-1">'
            f'<a href="#" class="alert-link text-primary fw-semibold" '
            f'hx-get="/dict/objects/{obj.uuid}/?sidebar=1" '
            f'hx-target="#detail-container" '
            f'hx-on:click="bootstrap.Modal.getInstance(document.getElementById(\'createObjectModal\')).hide();">'
            f'{obj.name or obj.model.name} <span class="text-muted fw-normal">({obj.model.name})</span>'
            f'</a></li>'
            for obj in similar_objects
        ]
        return HttpResponse(
            f'<div class="alert alert-warning py-2 px-3 mt-2 mb-0 rounded-3 small animate-fade">'
            f'<div class="fw-bold text-dark mb-1"><i class="bi bi-exclamation-triangle-fill text-warning me-1"></i> Похожие объекты ({len(similar_objects)} шт.):</div>'
            f'<ul class="ps-3 mb-1" style="max-height: 80px; overflow-y: auto;">{"".join(links)}</ul>'
            f'</div>'
        )

    return HttpResponse('<div class="text-success small mt-1"><i class="bi bi-check-circle-fill me-1"></i> Имя свободно</div>')


@login_required
def suggest_view(request):
    field = request.GET.get('field')
    q = request.GET.get('q', '').strip()
    
    if not q or len(q) < 1:
        return HttpResponse('')
        
    results = []
    words = q.split()
    
    if field == 'object_type':
        exact = ObjectType.objects.filter(type__iexact=q)
        word_filter = Q()
        for w in words:
            word_filter &= Q(type__icontains=w)
        partial = ObjectType.objects.filter(word_filter).exclude(pk__in=exact)
        results = list(exact) + list(partial)
        
    elif field == 'model':
        exact = ObjectModel.objects.filter(name__iexact=q)
        word_filter = Q()
        for w in words:
            word_filter &= (Q(name__icontains=w) | Q(object_type__type__icontains=w))
        partial = ObjectModel.objects.filter(word_filter).exclude(pk__in=exact)
        results = list(exact) + list(partial)
        
    # Поддерживаем и выбор родителя, и выбор исходного объекта для копирования
    elif field in ['parent', 'source_object']:
        exact = DataObject.objects.filter(Q(name__iexact=q) | Q(inventory_number__iexact=q))
        word_filter = Q()
        for w in words:
            word_filter &= (Q(name__icontains=w) | Q(inventory_number__icontains=w) | Q(model__name__icontains=w))
        partial = DataObject.objects.filter(word_filter).exclude(pk__in=exact)
        results = list(exact) + list(partial)

    elif field == 'date_update_rule':
        exact = DateUpdateRule.objects.filter(name__iexact=q)
        word_filter = Q()
        for w in words:
            word_filter &= Q(name__icontains=w)
        partial = DateUpdateRule.objects.filter(word_filter).exclude(pk__in=exact)
        results = list(exact) + list(partial)

    show_create_option = False
    if field in ['object_type', 'date_update_rule']:
        has_exact_match = any(
            (getattr(item, 'type' if field == 'object_type' else 'name', '').lower() == q.lower()) for item in results
        )
        if not has_exact_match:
            show_create_option = True

    return render(request, 'data/includes/suggestions_list.html', {
        'results': results[:10],
        'field': field,
        'q': q,
        'show_create_option': show_create_option
    })


@login_required
def select_suggestion_view(request):
    field = request.GET.get('field')
    uuid_val = request.GET.get('uuid')
    name_val = request.GET.get('name')
    
    display_name = ""
    hidden_name = field
    hidden_value = ""
    
    if uuid_val:
        hidden_value = uuid_val
        if field == 'object_type':
            display_name = get_object_or_404(ObjectType, pk=uuid_val).type
        elif field == 'model':
            model_obj = get_object_or_404(ObjectModel, pk=uuid_val)
            display_name = f"{model_obj.name} ({model_obj.object_type.type})"
        elif field in ['parent', 'source_object']:
            obj = get_object_or_404(DataObject, pk=uuid_val)
            display_name = obj.name or obj.model.name
        elif field == 'date_update_rule':
            display_name = get_object_or_404(DateUpdateRule, pk=uuid_val).name
    elif name_val:
        display_name = f"{name_val} (Создать новое)"
        hidden_name = f"new_{field}"
        hidden_value = name_val
        
    return render(request, 'data/includes/suggestion_selected.html', {
        'field': field,
        'display_name': display_name,
        'hidden_name': hidden_name,
        'hidden_value': hidden_value
    })


@login_required
def reset_suggestion_view(request):
    field = request.GET.get('field')
    placeholders = {
        'object_type': 'Введите тип оборудования...',
        'model': 'Введите модель оборудования...',
        'parent': 'Поиск родительского объекта...',
        'source_object': 'Поиск объекта для копирования...',
        'date_update_rule': 'Введите правило обновления...'
    }
    return render(request, 'data/includes/suggestion_input.html', {
        'field': field,
        'placeholder': placeholders.get(field, 'Начните вводить...')
    })


@login_required
def specs_builder_view(request):
    keys = request.POST.getlist('spec_keys')
    values = request.POST.getlist('spec_values')
    specs = dict(zip(keys, values))
    
    new_key = request.POST.get('new_key', '').strip()
    new_value = request.POST.get('new_value', '').strip()
    if new_key and new_value:
        specs[new_key] = new_value
        
    remove_key = request.POST.get('remove_key')
    if remove_key:
        specs.pop(remove_key, None)
        
    return render(request, 'data/includes/specs_builder.html', {
        'specifications': specs
    })


# --- КОНСТРУКТОР ПРАВИЛ И РАСЧЕТ СРОКОВ ТО ---

def calculate_next_maintenance_date(data_object, base_date=None):
    """Рассчитывает следующую дату ТО типа datetime.date"""
    if not data_object.date_update_rule:
        return None
        
    rule_data = data_object.date_update_rule.rule or {}
    strategy = rule_data.get('strategy', 'relative')
    anchor_type = rule_data.get('anchor', 'actual')
    value = rule_data.get('value', {})
    
    if not base_date:
        base_date = timezone.localdate()
    elif isinstance(base_date, datetime):
        base_date = base_date.date()
        
    if anchor_type == 'scheduled' and data_object.next_maintenance_date:
        base_date = data_object.next_maintenance_date

    if strategy == 'relative':
        delta = relativedelta(
            years=int(value.get('years', 0)),
            months=int(value.get('months', 0)),
            days=int(value.get('days', 0))
        )
        return base_date + delta
        
    elif strategy == 'fixed':
        if not isinstance(value, list) or not value:
            return None
            
        dates_in_year = []
        for item in value:
            m = int(item.get('month', 1))
            d = int(item.get('day', 1))
            try:
                dates_in_year.append(base_date.replace(month=m, day=d))
            except ValueError:
                dates_in_year.append(base_date.replace(month=m, day=28))
                
        dates_in_year.sort()
        
        for candidate in dates_in_year:
            if candidate > base_date:
                return candidate
                
        first_candidate = dates_in_year[0]
        return first_candidate + relativedelta(years=1)
        
    return None


@login_required
@role_required(['senior', 'admin', 'superuser'])
def rules_dates_builder_view(request):
    months = request.POST.getlist('fixed_months')
    days = request.POST.getlist('fixed_days')
    
    dates = []
    for m, d in zip(months, days):
        try:
            dates.append({'month': int(m), 'day': int(d)})
        except (ValueError, TypeError):
            pass
        
    new_month = request.POST.get('new_fixed_month')
    new_day = request.POST.get('new_fixed_day')
    
    if new_month and new_day:
        try:
            m_int = max(1, min(int(new_month), 12))  # Месяц строго от 1 до 12
            d_int = max(1, min(int(new_day), 31))    # День строго от 1 до 31
            new_date = {'month': m_int, 'day': d_int}
            if new_date not in dates:
                dates.append(new_date)
        except (ValueError, TypeError):
            pass
            
    remove_idx = request.POST.get('remove_idx')
    if remove_idx is not None:
        try:
            dates.pop(int(remove_idx))
        except (IndexError, ValueError):
            pass
            
    dates.sort(key=lambda x: (x['month'], x['day']))
    
    month_names = {
        1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля',
        5: 'Мая', 6: 'Июня', 7: 'Июля', 8: 'Августа',
        9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
    }
    
    return render(request, 'data/includes/rules_dates_builder.html', {
        'dates': dates,
        'month_names': month_names
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def rule_constructor_view(request):
    strategy = request.GET.get('new_rule_strategy', 'relative')
    anchor = request.GET.get('new_rule_anchor', 'actual')
    context = {
        'strategy': strategy,
        'anchor': anchor,
        'years': 0,
        'months': 6,
        'days': 0,
        'fixed_dates': [],
    }
    if strategy == 'fixed':
        context['month_names'] = {
            1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля',
            5: 'Мая', 6: 'Июня', 7: 'Июля', 8: 'Августа',
            9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
        }
    return render(request, 'data/includes/rule_constructor_fields.html', context)


@login_required
@role_required(['senior', 'admin', 'superuser'])
def toggle_scheduling_mode_view(request):
    mode = request.GET.get('maintenance_scheduling_mode', 'manual')
    is_inline = request.GET.get('inline', '0') == '1'
    template = 'data/includes/scheduling_mode_fields_inline.html' if is_inline else 'data/includes/scheduling_mode_fields.html'
    return render(request, template, {'mode': mode})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_rule_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'POST':
        mode = request.POST.get('maintenance_scheduling_mode', 'manual')
        
        if mode == 'manual':
            obj.date_update_rule = None
            obj.save(update_fields=['date_update_rule'])
        else:
            rule_uuid = request.POST.get('date_update_rule')
            new_rule_name = request.POST.get('new_date_update_rule')
            
            selected_rule = None
            if rule_uuid:
                selected_rule = get_object_or_404(DateUpdateRule, pk=rule_uuid)
            elif new_rule_name:
                strategy = request.POST.get('new_rule_strategy', 'relative')
                if strategy == 'relative':
                    anchor = request.POST.get('new_rule_anchor', 'actual')
                    years = int(request.POST.get('new_rule_years', 0))
                    months = int(request.POST.get('new_rule_months', 6))
                    days = int(request.POST.get('new_rule_days', 0))
                    rule_json = {
                        "strategy": "relative",
                        "anchor": anchor,
                        "value": {"years": years, "months": months, "days": days}
                    }
                elif strategy == 'fixed':
                    fixed_months = request.POST.getlist('fixed_months')
                    fixed_days = request.POST.getlist('fixed_days')
                    dates_list = []
                    for m, d in zip(fixed_months, fixed_days):
                        dates_list.append({"month": int(m), "day": int(d)})
                    rule_json = {"strategy": "fixed", "anchor": "yearly", "value": dates_list}
                    
                selected_rule, _ = DateUpdateRule.objects.get_or_create(
                    name=new_rule_name,
                    defaults={"rule": rule_json}
                )
                
            if selected_rule:
                obj.date_update_rule = selected_rule
                obj.next_maintenance_date = calculate_next_maintenance_date(obj, base_date=timezone.localdate())
                obj.save(update_fields=['date_update_rule', 'next_maintenance_date'])
                
        rule_name = obj.date_update_rule.name if obj.date_update_rule else 'ручной ввод'
        ActionHistory.objects.create(
            user=request.user,
            data_object=obj,
            action_type='rule_change', 
            action=f"Изменено правило планирования ТО на: {rule_name}."
        )
        
        rule_name_display = f'<span class="text-muted me-1">Правило ТО:</span><strong class="text-dark fw-semibold">{obj.date_update_rule.name}</strong><i class="bi bi-pencil-square edit-icon"></i>' if obj.date_update_rule else '<span class="text-muted me-1">Правило ТО:</span><strong class="text-dark fw-semibold">ручной ввод</strong><i class="bi bi-pencil-square edit-icon"></i>'
        
        oob_rule_html = f"""
        <span id="rule-display-container" 
              data-bs-toggle="modal" 
              data-bs-target="#editRuleModal"
              hx-get="/dict/objects/{obj.uuid}/edit-rule/" 
              hx-target="#edit-rule-modal-content"
              style="cursor: pointer;" 
              class="transition-all editable-trigger"
              hx-swap-oob="outerHTML">
            {rule_name_display}
        </span>
        """
        
        date_str = obj.next_maintenance_date.strftime('%d.%m.%Y') if obj.next_maintenance_date else "Не запланировано"
        oob_date_html = f'<strong id="maintenance-date-display" class="text-danger" hx-swap-oob="innerHTML">{date_str}</strong>'
        
        return HttpResponse(f"{oob_rule_html}\n{oob_date_html}")

    current_mode = 'auto' if obj.date_update_rule else 'manual'
    rule_details = None
    
    if obj.date_update_rule:
        rule_data = obj.date_update_rule.rule or {}
        strategy = rule_data.get('strategy', 'relative')
        
        if strategy == 'relative':
            anchor_text = 'фактического выполнения' if rule_data.get('anchor') == 'actual' else 'планового срока'
            val = rule_data.get('value', {})
            rule_details = f"Интервал от {anchor_text}: {val.get('years', 0)}г. {val.get('months', 0)}мес. {val.get('days', 0)}дн."
        elif strategy == 'fixed':
            month_names = {
                1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
                7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
            }
            dates = rule_data.get('value', [])
            formatted_dates = [f"{d.get('day')} {month_names.get(d.get('month'))}" for d in dates]
            rule_details = f"Сезонные даты: {', '.join(formatted_dates)}"

    return render(request, 'data/object/edit_rule_modal_body.html', {
        'obj': obj,
        'current_mode': current_mode,
        'rule_details': rule_details
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_object_model_view(request, pk):
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'GET' and request.GET.get('cancel') == '1':
        return render(request, 'data/object/inline_model.html', {'obj': obj, 'editing': False})
        
    if request.method == 'POST':
        model_uuid = request.POST.get('model')
        
        if model_uuid:
            new_model = get_object_or_404(ObjectModel, pk=model_uuid)
            old_model = obj.model
            
            if old_model != new_model:
                obj.model = new_model
                obj.save(update_fields=['model'])
                
                ActionHistory.objects.create(
                    user=request.user,
                    data_object=obj,
                    action_type='update',
                    action=f"Модель оборудования изменена с '{old_model.name}' на '{new_model.name}'."
                )
                
        model_html = render_to_string('data/object/inline_model.html', {'obj': obj, 'editing': False}, request=request)
        sidebar_node_html = render_to_string('data/tree/object_tree_node_label.html', {
            'node': obj,
            'is_active': True,
            'oob': True
        }, request=request)
        
        return HttpResponse(model_html + "\n" + sidebar_node_html)
        
    return render(request, 'data/object/inline_model.html', {'obj': obj, 'editing': True})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def edit_model_name_view(request, pk):
    model_obj = get_object_or_404(ObjectModel, pk=pk)
    
    if request.method == 'GET' and request.GET.get('cancel') == '1':
        return render(request, 'data/model/inline_model_name.html', {'model_obj': model_obj, 'editing': False})
        
    if request.method == 'POST':
        old_name = model_obj.name
        new_name = request.POST.get('name', '').strip()
        
        if new_name and old_name != new_name:
            model_obj.name = new_name
            model_obj.save(update_fields=['name'])
            
        model_name_html = render_to_string('data/model/inline_model_name.html', {'model_obj': model_obj, 'editing': False}, request=request)
        sidebar_model_node_html = render_to_string('data/tree/model_tree_node_label.html', {
            'model': model_obj,
            'is_active': True,
            'oob': True
        }, request=request)
        
        return HttpResponse(model_name_html + "\n" + sidebar_model_node_html)
        
    return render(request, 'data/model/inline_model_name.html', {'model_obj': model_obj, 'editing': True})


@login_required
@role_required(['senior', 'admin', 'superuser'])
def set_preview_attachment_view(request, pk):
    attachment = get_object_or_404(Attachment.objects.select_related('data_object'), pk=pk)
    obj = attachment.data_object
    
    if not attachment.is_image:
        return HttpResponse("Только изображение может быть превью", status=400)
    
    if attachment.is_preview:
        attachment.is_preview = False
        attachment.save(update_fields=['is_preview'])
    else:
        Attachment.objects.filter(data_object=obj, is_preview=True).update(is_preview=False)
        attachment.is_preview = True
        attachment.save(update_fields=['is_preview'])

    files = obj.attachments.select_related('user', 'comment').order_by('-created_at')
    return render(request, 'data/object/object_tab_files.html', {'obj': obj, 'files': files})


@login_required
def sync_youtrack_view(request, pk):
    """Запуск синхронизации с YouTrack"""
    obj = get_object_or_404(DataObject, pk=pk)
    
    if request.method == 'POST':
        sync_issue_from_youtrack(obj, request.user)
        return object_detail_view(request, pk)

    return HttpResponse("Метод не разрешен", status=405)


# --- СЕРВИС И ПРЕДСТАВЛЕНИЯ КЛОНИРОВАНИЯ ОБЪЕКТОВ ---

def deep_clone_object(source_obj, new_root_name, new_parent=None, source_root_name=None, user=None, clone_children=True):
    """
    Рекурсивно клонирует объект и всю его дочернюю иерархию с умным суффиксированием:
    - Если в названии детали было имя старого родителя -> подменяем на новое имя.
    - Если название детали общее (например, "Блок питания") -> добавляем суффикс "(НовоеИмя)".
    """
    is_root = (source_root_name is None)
    if is_root:
        source_root_name = source_obj.name or (source_obj.model.name if source_obj.model else "")
    
    # 1. Формируем имя для текущего узла
    if is_root:
        # Это сам корневой объект
        obj_name = new_root_name
    else:
        # Это дочерняя деталь
        orig_name = source_obj.name or (source_obj.model.name if source_obj.model else "Компонент")
        if source_root_name and source_root_name in orig_name:
            # Заменяем старое имя родителя на новое
            obj_name = orig_name.replace(source_root_name, new_root_name)
        else:
            # Если имя общее — приписываем суффикс нового родителя
            obj_name = f"{orig_name} ({new_root_name})"

    # 2. Создаем копию объекта
    cloned_obj = DataObject.objects.create(
        name=obj_name,
        model=source_obj.model,
        parent=new_parent,
        inventory_number=None,       # Очищаем инвентарник
        youtrack_issue_id=None,      # Очищаем задачу YouTrack
        next_maintenance_date=source_obj.next_maintenance_date,
        date_update_rule=source_obj.date_update_rule,
        description=source_obj.description
    )
    
    # 3. Фиксируем создание в истории объекта
    ActionHistory.objects.create(
        user=user,
        data_object=cloned_obj,
        action_type='create',
        action=f"Объект создан копированием из '{source_obj.name or source_obj.model.name}'."
    )
    
    # 4. Рекурсивно клонируем всех потомков
    if clone_children:
        for child in source_obj.children.all().order_by('name'):
            deep_clone_object(
                source_obj=child,
                new_root_name=new_root_name,
                new_parent=cloned_obj,
                source_root_name=source_root_name,
                user=user,
                clone_children=True
            )
            
    return cloned_obj



@login_required
@role_required(['senior', 'admin', 'superuser'])
def clone_object_modal_view(request):
    """Отображение модального окна тиражирования объекта"""
    active_object_id = request.GET.get('active_object_id') or request.session.get('active_object_id')
    preselected_object = None
    if active_object_id:
        try:
            preselected_object = DataObject.objects.get(pk=active_object_id)
        except DataObject.DoesNotExist:
            pass

    return render(request, 'data/includes/clone_object_modal_body.html', {
        'preselected_object': preselected_object
    })


@login_required
@role_required(['senior', 'admin', 'superuser'])
def clone_object_view(request):
    """Создание копии объекта (обычной или с дочерними элементами)"""
    if request.method == 'POST':
        # Поддерживаем оба имени параметра: source_object и parent (из подсказок поиска)
        source_uuid = request.POST.get('source_object') or request.POST.get('parent')
        new_name = request.POST.get('new_name', '').strip()
        clone_children = request.POST.get('clone_children') == 'on'
        keep_parent = request.POST.get('keep_parent') == 'on'
        
        if source_uuid:
            source_obj = get_object_or_404(DataObject.objects.select_related('parent'), pk=source_uuid)
            
            if not new_name:
                new_name = f"{source_obj.name or source_obj.model.name} (копия)"
                
            new_parent = source_obj.parent if keep_parent else None
            
            deep_clone_object(
                source_obj=source_obj,
                new_root_name=new_name,
                new_parent=new_parent,
                user=request.user,
                clone_children=clone_children
            )

        roots = DataObject.objects.filter(parent__isnull=True).prefetch_related('children').order_by('name')
        context = {
            'initial_objects': roots,
            'active_tab': 'objects',
            'models': ObjectModel.objects.all().order_by('name'),
            'object_types': ObjectType.objects.all().order_by('type')
        }
        return render(request, 'data/tree/dict_sidebar.html', context)

    return HttpResponse("Метод не разрешен", status=405)